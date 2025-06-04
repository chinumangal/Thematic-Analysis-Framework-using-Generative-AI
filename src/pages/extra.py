from openpyxl import load_workbook
import os, time

local_dir: str = os.path.abspath(os.path.join(__file__ ,"../../../data/"))
output_file = os.path.join(local_dir, "views", "view_implications.xlsx")


if os.path.exists(output_file):
    wb = load_workbook(output_file)
    if "Domain Implications" in wb.sheetnames:
        std = wb["Domain Implications"]
        wb.remove(std)
        wb.save(output_file)